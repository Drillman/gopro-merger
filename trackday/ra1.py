"""Lecteur des fichiers .ra1 du chronometre GPS 3DMS.

Format (little-endian, reconstruit par retro-ingenierie) :

    [pstr] magic       "RA1"        (octet de longueur + ascii)
    [pstr] version     "1.0.0.0"
    u32    count       nombre d'echantillons
    count * 28 octets  echantillons :
        i32   t          horodatage relatif en ms (pas fixe de 100 ms -> 10 Hz)
        f32   lon        longitude WGS84 (degres)
        f32   lat        latitude WGS84 (degres)
        f32   speed      vitesse sol (km/h)
        f32   _pad0      toujours 0 dans les fichiers observes
        f32   accel      acceleration longitudinale (g), pas de 2/127, sature a +/-2
        f32   _pad1      toujours 0 dans les fichiers observes
    25 octets          bloc inconnu, nul sauf l'octet 15 (=1 quand le bloc
                       final ci-dessous est absent)
    bloc final (present seulement si le fichier se termine par AB CD EF) :
        [pstr] circuit         ex. "Ferte Gaucher"  (utf-8)
        2 octets nuls
        [pstr] best_lap        ex. "02:07.55"
        [pstr] theoretical     ex. "02:06.51"  (tour theorique, somme des meilleurs secteurs)
        1 octet nul
        [pstr] conditions      ex. "Dry"
        u32    track_id        20 sur tous les fichiers observes
        3 octets  AB CD EF     marqueur de fin

Le fichier ne contient aucune date absolue : elle est portee par le nom du
fichier ("AAAA-MM-JJ a HHhMM"), qui correspond a l'heure de FIN de session.

Il ne contient pas non plus le decoupage en secteurs du boitier : les secteurs
sont donc redefinis ici a distance egale le long du meilleur tour (voir
--secteurs), ce qui rend le "temps ideal" comparable d'une session a l'autre
mais pas identique au "theorique" affiche par le 3DMS.

Usage : python ra1.py                      -> analyse dump/GPS, classeur par jour
        python ra1.py "dump/GPS/2026-07-26 a 09h19.ra1"
        python ra1.py dump/GPS --out output/GPS --secteurs 3
"""

import argparse
import csv
import math
import struct
import sys
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from pathlib import Path

BASE = Path(__file__).resolve().parent.parent
REC_SIZE = 28
REC_FMT = "<i6f"
TAIL_MAGIC = b"\xab\xcd\xef"
NAME_FMT = "%Y-%m-%d a %Hh%M"

HALF_WIDTH = 30.0   # demi-largeur des lignes de detection, en metres
MIN_SPEED = 30.0    # km/h, evite de valider un franchissement a l'arret


@dataclass
class Sample:
    t: float        # secondes depuis le debut de session
    lat: float
    lon: float
    speed: float    # km/h
    accel: float    # g, longitudinal


@dataclass
class Session:
    path: Path
    version: str
    samples: list
    circuit: str = ""
    best_lap: str = ""
    theoretical: str = ""
    conditions: str = ""
    track_id: int = 0
    finalized: bool = False
    _xy: list = field(default=None, repr=False)

    @property
    def duration(self) -> float:
        return self.samples[-1].t - self.samples[0].t if self.samples else 0.0

    @property
    def end_time(self):
        """Heure de fin de session, deduite du nom de fichier (heure locale)."""
        try:
            return datetime.strptime(self.path.stem, NAME_FMT)
        except ValueError:
            return None

    @property
    def start_time(self):
        end = self.end_time
        return end - timedelta(seconds=self.duration) if end else None

    @property
    def xy(self):
        """Projection plane metrique locale, centree sur le premier point."""
        if self._xy is None:
            lat0, lon0 = self.samples[0].lat, self.samples[0].lon
            kx = 111320.0 * math.cos(math.radians(lat0))
            ky = 110540.0
            self._xy = [((s.lon - lon0) * kx, (s.lat - lat0) * ky) for s in self.samples]
        return self._xy


@dataclass
class Line:
    """Ligne de detection : un point et la normale (= sens de marche)."""
    px: float
    py: float
    hx: float
    hy: float

    def ahead(self, p):
        return (p[0] - self.px) * self.hx + (p[1] - self.py) * self.hy

    def offset(self, p):
        return abs(-(p[0] - self.px) * self.hy + (p[1] - self.py) * self.hx)


# --------------------------------------------------------------------------- lecture

def _pstr(buf, off):
    n = buf[off]
    return buf[off + 1:off + 1 + n].decode("utf-8"), off + 1 + n


def load(path) -> Session:
    path = Path(path)
    d = path.read_bytes()

    magic, off = _pstr(d, 0)
    if magic != "RA1":
        raise ValueError(f"{path.name} : magic inattendu {magic!r}")
    version, off = _pstr(d, off)
    count, = struct.unpack_from("<I", d, off)
    off += 4

    expected = off + count * REC_SIZE
    if len(d) < expected:
        raise ValueError(f"{path.name} : fichier tronque "
                         f"({len(d)} octets, {expected} attendus)")

    samples = []
    for i in range(count):
        t, lon, lat, speed, _, accel, _ = struct.unpack_from(REC_FMT, d, off + i * REC_SIZE)
        samples.append(Sample(t / 1000.0, lat, lon, speed, accel))

    s = Session(path=path, version=version, samples=samples)

    tail = d[expected:]
    if tail.endswith(TAIL_MAGIC):
        o = 25
        s.circuit, o = _pstr(tail, o)
        o += 2
        s.best_lap, o = _pstr(tail, o)
        s.theoretical, o = _pstr(tail, o)
        o += 1
        s.conditions, o = _pstr(tail, o)
        s.track_id, = struct.unpack_from("<I", tail, o)
        s.finalized = True
    return s


# --------------------------------------------------------------------------- geometrie

def distance(session, i0=0, i1=None) -> float:
    """Distance parcourue en metres entre deux indices d'echantillon."""
    xy = session.xy
    i1 = len(xy) - 1 if i1 is None else i1
    return sum(math.dist(xy[i - 1], xy[i]) for i in range(i0 + 1, i1 + 1))


def finish_line(session) -> Line:
    """Le chrono demarre sur la ligne : t=0 est le franchissement, la position
    de la ligne est donc extrapolee en arriere depuis les deux premiers points."""
    sm, xy = session.samples, session.xy
    f = sm[0].t / (sm[1].t - sm[0].t)
    px = xy[0][0] - f * (xy[1][0] - xy[0][0])
    py = xy[0][1] - f * (xy[1][1] - xy[0][1])
    hx, hy = xy[1][0] - xy[0][0], xy[1][1] - xy[0][1]
    n = math.hypot(hx, hy)
    return Line(px, py, hx / n, hy / n)


def crossings(session, line, t0=None, t1=None, first_only=False):
    """Instants de franchissement de `line`, interpoles, dans [t0, t1]."""
    sm, xy = session.samples, session.xy
    out = []
    prev = None
    for i, s in enumerate(sm):
        if t0 is not None and s.t < t0:
            continue
        if t1 is not None and s.t > t1:
            break
        a = line.ahead(xy[i])
        if prev is not None and prev[1] < 0 <= a \
                and line.offset(xy[i]) < HALF_WIDTH and s.speed > MIN_SPEED:
            r = -prev[1] / (a - prev[1])
            out.append(prev[0] + r * (s.t - prev[0]))
            if first_only:
                return out
        prev = (s.t, a)
    return out


def laps(session):
    """(instants de franchissement de la ligne, temps au tour). Le premier
    franchissement est t=0, le depart du chrono."""
    if len(session.samples) < 3:
        return [], []
    c = [0.0] + crossings(session, finish_line(session))
    return c, [c[i + 1] - c[i] for i in range(len(c) - 1)]


def flying_laps(session, tolerance=0.05):
    """Indices des tours representatifs, hors sortie et rentree des stands.

    Le chrono demarre au franchissement de la ligne en sortant des stands : le
    premier tour est donc toujours un tour de mise en temperature, et le dernier
    peut etre un tour de rentree. On retire les tours de DEBUT et de FIN dont le
    temps depasse le meilleur de plus de `tolerance`.

    On ne rogne que les extremites : un tour lent en milieu de session (trafic,
    drapeau jaune) reste un tour roule et doit etre conserve. Le meilleur tour
    etant par construction sous le seuil, la liste ne peut pas se vider.
    """
    cross, times = laps(session)
    if not times:
        return []
    limit = min(times) * (1 + tolerance)
    keep = list(range(len(times)))
    while keep and times[keep[0]] > limit:
        keep.pop(0)
    while keep and times[keep[-1]] > limit:
        keep.pop()
    return keep


def sector_lines(session, n, cross, lap_times):
    """n-1 lignes intermediaires, a distance egale le long du meilleur tour.

    Les lignes sont des positions fixes sur la piste, donc identiques pour tous
    les tours : les temps par secteur restent comparables meme si la trajectoire
    varie d'un tour a l'autre.
    """
    if n < 2 or not lap_times:
        return []
    k = lap_times.index(min(lap_times))
    t0, t1 = cross[k], cross[k + 1]
    sm, xy = session.samples, session.xy
    idx = [i for i, s in enumerate(sm) if t0 <= s.t <= t1]
    if len(idx) < n + 2:
        return []

    cum = [0.0]
    for a, b in zip(idx, idx[1:]):
        cum.append(cum[-1] + math.dist(xy[a], xy[b]))
    total = cum[-1]

    lines = []
    for j in range(1, n):
        target = total * j / n
        m = next((q for q in range(1, len(cum)) if cum[q] >= target), len(cum) - 1)
        seg = cum[m] - cum[m - 1]
        r = (target - cum[m - 1]) / seg if seg else 0.0
        a, b = xy[idx[m - 1]], xy[idx[m]]
        px, py = a[0] + r * (b[0] - a[0]), a[1] + r * (b[1] - a[1])
        hx, hy = b[0] - a[0], b[1] - a[1]
        h = math.hypot(hx, hy)
        lines.append(Line(px, py, hx / h, hy / h))
    return lines


@dataclass
class Lap:
    num: int
    t_start: float
    total: float
    sectors: list       # temps par secteur, None si la ligne n'a pas ete franchie
    vmax: float


def lap_table(session, n_sectors=3):
    """Detail de chaque tour : temps par secteur, temps total, vitesse max."""
    cross, lap_times = laps(session)
    if not lap_times:
        return []
    lines = sector_lines(session, n_sectors, cross, lap_times)

    table = []
    for k, total in enumerate(lap_times):
        t0, t1 = cross[k], cross[k + 1]
        bounds = [t0]
        for ln in lines:
            c = crossings(session, ln, t0=bounds[-1], t1=t1, first_only=True)
            if not c:
                bounds = None
                break
            bounds.append(c[0])
        if bounds is None:
            sectors = [None] * n_sectors
        else:
            bounds.append(t1)
            sectors = [bounds[j + 1] - bounds[j] for j in range(n_sectors)]
        vmax = max((s.speed for s in session.samples if t0 <= s.t <= t1), default=0.0)
        table.append(Lap(k + 1, t0, total, sectors, vmax))
    return table


# --------------------------------------------------------------------------- sorties

def fmt_lap(sec) -> str:
    return "" if sec is None else f"{int(sec // 60)}:{sec % 60:05.2f}"


def to_csv_points(session, dest: Path):
    """Export brut : un echantillon par ligne (pour les overlays video)."""
    cross, _ = laps(session)
    with dest.open("w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["t", "lap", "t_lap", "lat", "lon", "speed_kmh", "accel_g"])
        for s in session.samples:
            lap = sum(1 for c in cross if c <= s.t)
            t_lap = s.t - cross[lap - 1] if lap else 0.0
            w.writerow([f"{s.t:.1f}", lap, f"{t_lap:.1f}",
                        f"{s.lat:.7f}", f"{s.lon:.7f}",
                        f"{s.speed:.2f}", f"{s.accel:.5f}"])


def _sheet(wb, session, n_sectors):
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    DUR = "m:ss.00"
    GAP = '+0.00" s";-0.00" s";"—"'
    ARIAL = "Arial"
    head_fill = PatternFill("solid", fgColor="1F3864")
    best_fill = PatternFill("solid", fgColor="E2EFDA")
    thin = Side(style="thin", color="BFBFBF")

    title = session.end_time.strftime("%Hh%M") if session.end_time else session.path.stem
    ws = wb.create_sheet(title[:31])
    table = lap_table(session, n_sectors)

    ws["A1"] = f"{session.circuit or 'Circuit inconnu'} — session {title}"
    ws["A1"].font = Font(ARIAL, size=14, bold=True)
    start, end = session.start_time, session.end_time
    infos = []
    if start:
        infos.append(f"{start:%d/%m/%Y} · {start:%H:%M:%S} → {end:%H:%M:%S}")
    if session.conditions:
        infos.append(session.conditions)
    infos += [f"{len(table)} tours",
              f"{distance(session) / 1000:.2f} km",
              f"vmax {max(s.speed for s in session.samples):.1f} km/h"]
    ws["A2"] = " · ".join(infos)
    ws["A2"].font = Font(ARIAL, size=10, color="595959")

    cols = ["Tour"] + [f"S{j + 1}" for j in range(n_sectors)] + ["Temps", "Écart", "Vmax km/h"]
    r0 = 4
    for c, name in enumerate(cols, 1):
        cell = ws.cell(r0, c, name)
        cell.font = Font(ARIAL, bold=True, color="FFFFFF")
        cell.fill = head_fill
        cell.alignment = Alignment(horizontal="center")

    first, last = r0 + 1, r0 + len(table)
    best_row = last + 2
    for i, lap in enumerate(table):
        r = first + i
        ws.cell(r, 1, lap.num).alignment = Alignment(horizontal="center")
        for j, sec in enumerate(lap.sectors):
            cell = ws.cell(r, 2 + j, sec / 86400 if sec is not None else None)
            cell.number_format = DUR
        tc = ws.cell(r, 2 + n_sectors, lap.total / 86400)
        tc.number_format = DUR
        tc.font = Font(ARIAL, bold=True)
        gc = ws.cell(r, 3 + n_sectors,
                     f"=({tc.coordinate}-${tc.column_letter}${best_row})*86400")
        gc.number_format = GAP
        ws.cell(r, 4 + n_sectors, round(lap.vmax, 1)).number_format = "0.0"

    def col(c):
        return ws.cell(r0, c).column_letter

    # meilleurs : par secteur et sur le tour complet
    ws.cell(best_row, 1, "Meilleur").font = Font(ARIAL, bold=True)
    for c in range(2, 3 + n_sectors):
        cell = ws.cell(best_row, c, f"=MIN({col(c)}{first}:{col(c)}{last})")
        cell.number_format = DUR
        cell.font = Font(ARIAL, bold=True)
        cell.fill = best_fill

    # tour ideal = somme des meilleurs secteurs
    ideal_row = best_row + 1
    ws.cell(ideal_row, 1, "Idéal").font = Font(ARIAL, bold=True)
    if n_sectors >= 2:
        cell = ws.cell(ideal_row, 2 + n_sectors,
                       f"=SUM({col(2)}{best_row}:{col(1 + n_sectors)}{best_row})")
        cell.number_format = DUR
        cell.font = Font(ARIAL, bold=True)
        cell.fill = best_fill

    for r in range(r0, ideal_row + 1):
        for c in range(1, 5 + n_sectors):
            ws.cell(r, c).border = Border(bottom=thin)

    note_row = ideal_row + 2
    if session.finalized:
        ws.cell(note_row, 1,
                f"Chrono 3DMS : meilleur {session.best_lap}, théorique {session.theoretical}")
    else:
        ws.cell(note_row, 1, "Chrono 3DMS : session non finalisée, pas de temps de référence")
    ws.cell(note_row, 1).font = Font(ARIAL, size=9, italic=True, color="595959")
    ws.cell(note_row + 1, 1,
            f"Idéal = somme des {n_sectors} meilleurs secteurs. Secteurs recalculés en "
            f"{n_sectors} portions de distance égale le long du meilleur tour "
            f"(le découpage du boîtier n'est pas stocké dans le .ra1)")
    ws.cell(note_row + 1, 1).font = Font(ARIAL, size=9, italic=True, color="595959")

    ws.column_dimensions["A"].width = 10
    for c in range(2, 5 + n_sectors):
        ws.column_dimensions[col(c)].width = 11
    ws.freeze_panes = ws.cell(first, 1)
    for row in ws.iter_rows():
        for cell in row:
            if cell.font.name != ARIAL:
                cell.font = Font(ARIAL, size=cell.font.size, bold=cell.font.bold,
                                 italic=cell.font.italic, color=cell.font.color)
    return ws


def to_xlsx(sessions, dest: Path, n_sectors=3):
    """Un classeur, un onglet par session."""
    from openpyxl import Workbook

    wb = Workbook()
    wb.remove(wb.active)
    for s in sessions:
        _sheet(wb, s, n_sectors)
    wb.save(dest)


def dump(session, n_sectors=3):
    s = session
    print(f"\n=== {s.path.name}  (RA1 v{s.version})")
    start, end = s.start_time, s.end_time
    if start:
        print(f"  session   : {start:%H:%M:%S} -> {end:%H:%M:%S} locale, {s.duration:.1f}s, "
              f"{len(s.samples)} points a 10 Hz")
    if s.finalized:
        print(f"  circuit   : {s.circuit} (id {s.track_id}), {s.conditions}")
        print(f"  chrono    : meilleur {s.best_lap}, theorique {s.theoretical}")
    else:
        print("  circuit   : bloc de metadonnees absent (session non finalisee)")
    print(f"  distance  : {distance(s) / 1000:.2f} km, "
          f"vmax {max(x.speed for x in s.samples):.1f} km/h")
    table = lap_table(s, n_sectors)
    if not table:
        return
    best = min(l.total for l in table)
    print(f"  {'tour':>5} " + " ".join(f"{'S' + str(j + 1):>8}" for j in range(n_sectors))
          + f" {'temps':>9} {'ecart':>7}")
    for l in table:
        gap = l.total - best
        print(f"  {l.num:>5} " + " ".join(f"{fmt_lap(x):>8}" for x in l.sectors)
              + f" {fmt_lap(l.total):>9} {('' if gap == 0 else f'+{gap:.2f}'):>7}")
    bests = [min((l.sectors[j] for l in table if l.sectors[j] is not None), default=None)
             for j in range(n_sectors)]
    print(f"  {'best':>5} " + " ".join(f"{fmt_lap(x):>8}" for x in bests)
          + f" {fmt_lap(best):>9}")
    if all(b is not None for b in bests):
        print(f"  {'ideal':>5} " + " " * (9 * n_sectors) + f"{fmt_lap(sum(bests)):>9}")


# --------------------------------------------------------------------------- cli

def main(argv=None):
    ap = argparse.ArgumentParser(prog="chrono",
                                 description="Lecteur des fichiers .ra1 du chrono GPS 3DMS.")
    ap.add_argument("target", nargs="?", default=BASE / "dump" / "GPS", type=Path,
                    help="fichier .ra1 ou dossier a analyser (defaut : dump/GPS)")
    ap.add_argument("--out", nargs="?", const=BASE / "output" / "GPS", default=None, type=Path,
                    metavar="DOSSIER",
                    help="ecrit un classeur .xlsx par journee (defaut : output/GPS)")
    ap.add_argument("--secteurs", type=int, default=3, metavar="N",
                    help="nombre de secteurs, a distance egale (defaut : 3)")
    ap.add_argument("--points", action="store_true",
                    help="exporte aussi le CSV brut 10 Hz de chaque session")
    a = ap.parse_args(argv)

    if a.secteurs < 1:
        ap.error("--secteurs doit valoir au moins 1")

    files = sorted(a.target.glob("*.ra1")) if a.target.is_dir() else [a.target]
    if not files:
        print(f"aucun .ra1 dans {a.target}")
        return 2

    if a.out:
        a.out.mkdir(parents=True, exist_ok=True)

    by_day = defaultdict(list)
    for f in files:
        s = load(f)
        dump(s, a.secteurs)
        if a.out and a.points:
            out = a.out / (f.stem + ".csv")
            to_csv_points(s, out)
            print(f"  -> {out}")
        end = s.end_time
        by_day[end.date() if end else "sans-date"].append(s)

    for day, sessions in sorted(by_day.items(), key=lambda kv: str(kv[0])):
        if not a.out:
            continue
        dest = a.out / f"{day}.xlsx"
        to_xlsx(sessions, dest, a.secteurs)
        print(f"\n{len(sessions)} session(s) -> {dest}")


if __name__ == "__main__":
    sys.exit(main() or 0)
